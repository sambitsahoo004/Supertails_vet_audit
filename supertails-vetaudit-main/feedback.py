import openpyxl
import pandas as pd
import re
from datetime import datetime
import os
import json
import argparse
import anthropic
from operator import itemgetter
from dotenv import load_dotenv
SUPPRESSED_IMPROVEMENT_PARAMETERS = [
    'previous_prescription',  # Suppress from improvements but keep in scoring
]

# Load environment variables from .env file
load_dotenv()

def extract_transcript_info(transcript_text):
    """Extract call information from the transcript header."""
    # Extract file name and duration
    file_info_match = re.search(
        r'File:\s*([^\s]+)\.mp3\s+\(Duration:\s+(\d+:\d+)\)', transcript_text)

    if file_info_match:
        full_file_name = file_info_match.group(1)
        duration = file_info_match.group(2)

        # Extract date and time from the file name
        date_match = re.search(r'_(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})$', full_file_name)
        if date_match:
            date_time_str = date_match.group(1)
            date_time_obj = datetime.strptime(date_time_str, '%Y-%m-%d_%H-%M-%S')
            formatted_date_time = date_time_obj.strftime('%Y-%m-%d %H:%M:%S')
        else:
            formatted_date_time = "Unknown"

        # Extract order ID (10-digit number)
        phone_match = re.search(r'(\d{10})', full_file_name)
        order_id = phone_match.group(1) if phone_match else "320"

        # Extract vet name from the filename before "_supertails"
        vet_name = "Unknown"
        try:
            # Remove trailing timestamp
            name_part = full_file_name.split("__")[0]
            if "_supertails" in name_part:
                raw_name = name_part.split("_supertails")[0]  # e.g., "bency_jhonson"
                name_tokens = raw_name.strip().split("_")
                if len(name_tokens) >= 2:
                    vet_name = f"Dr. {name_tokens[0].capitalize()} {name_tokens[1].capitalize()}"
                elif len(name_tokens) == 1:
                    vet_name = f"Dr. {name_tokens[0].capitalize()}"
        except Exception as e:
            vet_name = "Unknown"

        return {
            "order_id": order_id,
            "call_start_time": formatted_date_time,
            "call_duration": duration,
            "vet_name": vet_name
        }

    # Fallback values if match fails
    return {
        "order_id": "320",
        "call_start_time": "2025-05-04 00:00:00",
        "call_duration": "0:00",
        "vet_name": "Unknown"
    }


def find_lowest_scoring_parameter(category_data):
    """Find parameters with the lowest percentage scores in a category and their shortcomings."""
    lowest_params = []
    
    if 'parameter_scores' in category_data:
        params = []
        for param_name, param_data in category_data['parameter_scores'].items():
            # Include ALL parameters for clarification section (don't suppress any here)
            if 'percentage' in param_data:
                params.append((param_name, param_data))
        
        # Sort by percentage and get the lowest one
        if params:
            params.sort(key=lambda x: x[1]['percentage'])
            lowest_param = params[0]
            param_name = lowest_param[0].replace('_', ' ').title()
            shortcomings = lowest_param[1].get('shortcomings', 'No specific shortcomings identified')
            
            # Format as "Parameter: Shortcomings"
            lowest_params.append(f"{param_name}: {shortcomings}")
    
    return lowest_params

def find_lowest_scoring_parameter_for_feedback(category_data):
    """Find the parameter with the lowest percentage score in a category (excluding suppressed ones for improvements only)."""
    lowest_param = None
    lowest_score = float('inf')
    
    if 'parameter_scores' in category_data:
        for param_name, param_data in category_data['parameter_scores'].items():
            # Skip suppressed parameters only for feedback/improvements section
            if param_name.lower() not in SUPPRESSED_IMPROVEMENT_PARAMETERS:
                if 'percentage' in param_data and param_data['percentage'] < lowest_score:
                    lowest_score = param_data['percentage']
                    lowest_param = (param_name, param_data)
    
    return lowest_param

def extract_medicine_parameters(tech_assessment_data):
    """Extract medicine name and dosage parameters from technical assessment."""
    medicine_params = {}
    
    if 'parameter_scores' in tech_assessment_data:
        if 'medicine_name' in tech_assessment_data['parameter_scores']:
            medicine_params['medicine_name'] = tech_assessment_data['parameter_scores']['medicine_name']
        
        if 'medicine_dosage' in tech_assessment_data['parameter_scores']:
            medicine_params['medicine_dosage'] = tech_assessment_data['parameter_scores']['medicine_dosage']
    
    return medicine_params

def generate_feedback_summary(data):
    """Generate a doctor's feedback summary based on the data."""
    feedback_info = {}
    
    # Extract lowest scoring parameter from each category (except category 4)
    categories = [
        ('call_introduction', 'Call Introduction'),
        ('pet_information', 'Pet Information'),
        ('communication_quality', 'Communication Quality'),
        ('call_conclusion', 'Call Conclusion')
    ]
    
    for cat_key, cat_name in categories:
        if cat_key in data['categories']:
            category_data = data['categories'][cat_key]
            lowest_param = find_lowest_scoring_parameter_for_feedback(category_data)
            if lowest_param:
                param_name, param_data = lowest_param
                
                # Check if this parameter should be suppressed from improvements
                if param_name.lower() not in SUPPRESSED_IMPROVEMENT_PARAMETERS:
                    feedback_info[f"{cat_key}|{param_name}"] = {
                        'category_name': cat_name,
                        'param_name': param_name,
                        'score': param_data.get('percentage', 0),
                        'improvements': param_data.get('improvements', {}).get('recommendation', '') 
                                        if isinstance(param_data.get('improvements', {}), dict) 
                                        else param_data.get('improvements', [])
                    }
    
    # Extract medicine name and dosage from technical assessment
    if 'technical_assessment' in data['categories']:
        tech_assessment = data['categories']['technical_assessment']
        medicine_params = extract_medicine_parameters(tech_assessment)
        
        for param_name, param_data in medicine_params.items():
            display_name = param_name.replace('_', ' ').title()
            feedback_info[f"technical_assessment|{param_name}"] = {
                'category_name': 'Technical Assessment',
                'param_name': display_name,
                'score': param_data.get('percentage', 0),
                'improvements': param_data.get('improvements', '')
            }
    
    return feedback_info

def create_anthropic_prompt(feedback_info):
    """Create an improved prompt for the Anthropic API based on the feedback information."""
    prompt = """
    You are an expert veterinary consultant tasked with providing professional feedback to veterinarians based on their client call recordings. Your feedback should be insightful, actionable, and presented with clear, professional headings.

    ## YOUR TASK

    Analyze the assessment data below and create a structured feedback summary with clear, professional headings that reflect veterinary practice excellence.

    ## OUTPUT STRUCTURE

    - Start directly with the headings and bullet points
    - Do not add any introductory text, explanations, or conclusions
    - Produce ONLY the formatted feedback with no other content

    ## HEADING STYLE GUIDELINES

    - Create CLEAR, SIMPLE headings for each category that sound professional yet accessible
    - Focus on the key improvement area using straightforward language
    - Headings should be 3-5 words and immediately understandable
    - AVOID jargon, complex terminology, or overly academic phrasing
    - Each heading should clearly indicate the specific area needing improvement
    - GROUP similar improvements under a single appropriate heading

    ## HEADING EXAMPLES (Use these as inspiration, but create new ones)

    - "Client Introduction Improvement" (for call introduction issues)
    - "Pet Information Collection" (for missing pet information)
    - "Communication Tone Enhancement" (for tone problems)
    - "Effective Call Closing" (for call conclusion issues)
    - "Medication Advice Precision" (for medication dosage issues)
    - "Clear Treatment Explanation" (for unclear explanations)
    - "Aftercare Instruction Clarity" (for missing aftercare instructions)
    - "Complete History Taking" (for incomplete information gathering)
    - "Medication Communication Improvement" (for medication explanation issues)
    - "Confident Recommendation Delivery" (for uncertainty in recommendations)
    - "Follow-up Plan Creation" (for missing next steps)
    - "Client Concern Response" (for not addressing worries)
   
    ## OUTPUT FORMAT

    For each category (except Technical Assessment):

    Ø [CLEAR PROFESSIONAL HEADING]
    * [Group similar improvement points together, maintaining EXACT text from the data]
    * [Other improvement points from the same category]
   
    For Technical Assessment:

    Ø Technical Assessment -
    Medicine Name

    PRESCRIBED:
    - [medications mentioned - use EXACT text from data]

    AS PER STANDARD:
    - [ideal medications - use EXACT text from data]

    GAPS:
    - [gaps identified - use EXACT text from data]
   
    Medicine Dosage

    PRESCRIBED:
    - [dosages mentioned - use EXACT text from data]

    AS PER STANDARD:
    - [ideal dosages - use EXACT text from data]

    GAPS:
    - [gaps identified - use EXACT text from data]
   
    ## ASSESSMENT DATA
    """
    
    # Group by category for better prompt organization
    category_groups = {}
    tech_params = {}
    
    for key, info in feedback_info.items():
        cat_key, param_name = key.split('|')
        
        if cat_key == 'technical_assessment':
            tech_params[param_name] = info
        else:
            if cat_key not in category_groups:
                category_groups[cat_key] = []
            category_groups[cat_key].append((param_name, info))
    
    # Add regular categories to prompt
    for cat_key, params in category_groups.items():
        prompt += f"\n\nCategory: {params[0][1]['category_name']}\n"
        for param_name, info in params:
            readable_param = param_name.replace('_', ' ').title()
            prompt += f"Parameter: {readable_param} (Score: {info['score']}%)\n"
            
            # Handle different formats of improvements - preserve exact text
            if isinstance(info['improvements'], list):
                prompt += f"Improvements:\n"
                for imp in info['improvements']:
                    prompt += f"- {imp}\n"
            else:
                prompt += f"Improvements: {info['improvements']}\n"
    
    # Add technical assessment parameters if they exist
    if tech_params:
        prompt += "\n\nCategory: Technical Assessment\n"
        
        for param_name, info in tech_params.items():
            prompt += f"\nParameter: {info['param_name']} (Score: {info['score']}%)\n"
            prompt += f"Improvements: {info['improvements']}\n"
    
    prompt += """
    ## IMPORTANT REMINDERS

    1. Your feedback is analyzing a recorded CONVERSATION between a veterinarian and client - focus on communication aspects, professionalism, and clinical accuracy
    2. Create SIMPLE, CLEAR headings that anyone can understand at first glance
    3. Headings should be PROFESSIONAL but SIMPLE - avoid complex terminology or jargon
    4. For Technical Assessment, follow the exact structure provided
    5. Use the EXACT improvement points from the data - DO NOT modify or expand them in any way
    6. Always include the 'Ø' symbol before each heading
    7. IMPORTANT: Never paraphrase, modify, or enhance the improvement points - use them exactly as provided
    8. DO NOT use the word "Documentation" in your headings or bullet points

    ## CONTEXT

    Remember that this feedback is for a veterinarian after a client call - the headings should be easily understood by all veterinary staff including support personnel. Use clear, straightforward terminology that anyone in the veterinary practice can understand.

    The goal is to organize the exact feedback points under appropriate headings that are immediately clear to the reader.
    """
    
    return prompt
def get_anthropic_response(prompt):
    """Get a response from the Anthropic API."""
    try:
        # Get API key from environment variables
        api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key:
            # If no API key is available, return a sample feedback
            return generate_sample_feedback()
            
        client = anthropic.Anthropic(api_key=api_key)
        
        response = client.messages.create(
            model="claude-3-5-sonnet-latest",
            max_tokens=1000,
            temperature=0.5,  # Slightly increased for more creative heading generation
            system="You are an expert medical consultant providing structured feedback for veterinarians. Generate appropriate headings based on the parameters provided. Always include the 'Ø' symbol before each heading.",
            messages=[
                {"role": "user", "content": prompt}
            ]
        )
        
        return response.content[0].text
    except Exception as e:
        print(f"Error getting response from Anthropic API: {e}")
        # Return a sample feedback if API call fails
        return generate_sample_feedback()

def generate_sample_feedback():
    """Generate a sample feedback when Anthropic API is not available."""
    return """
Ø Streamlined Order Confirmation Process                                                        
* Verify order quantities and delivery preferences during call introduction to ensure accurate fulfillment

Ø Comprehensive Pet Profile Establishment
* Implement a structured checklist to collect complete demographic information, including pet name, age, and gender, for each patient

Ø Engaging Vocal Presence
* Incorporate more dynamic voice modulation and tone adjustments to enhance client engagement throughout the consultation

Ø Thorough Medication History Review
* Implement a standard protocol to review the pet's previous prescriptions during the call conclusion

Ø Technical Assessment -
Medicine Name

PRESCRIBED:
- ECPET deworming tablet, Brevicto 20-40 kg and 40-56 kg tablet, Praziplus

AS PER STANDARD:
- ECPET dosage based on Buddy's 25 kg weight
- Brevicto for flea/tick prevention every 3 months

GAPS:
- Praziplus vs. ECPET comparison not detailed
- Brevicto dosage not specified for Buddy's exact weight

Medicine Dosage

PRESCRIBED:
- ECPET deworming tablet, Brevicto 20-40 kg and 40-56 kg tablet

AS PER STANDARD:
- ECPET: 3 tablets for 25 kg
- Brevicto: 20-40 kg tablet every 3 months

GAPS:
- Brevicto dosage range needs adjustment for 25 kg dog
"""

def generate_feedback_for_excel(json_data):
    """Generate formatted feedback for Excel from JSON data."""
    # First generate the feedback information
    feedback_info = generate_feedback_summary(json_data)
    
    # Create the prompt for Anthropic
    prompt = create_anthropic_prompt(feedback_info)
    
    # Get response from Anthropic (or sample if API not available)
    feedback_summary = get_anthropic_response(prompt)
    
    # Format the feedback as a list for Excel
    formatted_feedback = []
    
    # Split the feedback by sections (using the Ø symbol)
    sections = feedback_summary.strip().split('Ø')
    
    # Process each section
    for section in sections:
        if not section.strip():
            continue
            
        # Add the section heading with Ø symbol
        lines = section.strip().split('\n')
        heading = lines[0].strip()
        formatted_feedback.append(f"Ø {heading}")
        
        # Add bullet points for this section - preserve original format
        for line in lines[1:]:
            line = line.strip()
            if line and line.startswith('*'):
                formatted_feedback.append(line)
            elif line and line.startswith('-'):
                formatted_feedback.append(line)
            elif line and line in ["PRESCRIBED:", "AS PER STANDARD:", "GAPS:"]:
                formatted_feedback.append(line)
            elif line and line in ["Medicine Name", "Medicine Dosage"]:
                formatted_feedback.append(line)
            elif line and not any(keyword in line.lower() for keyword in ['prescribed:', 'as per standard:', 'gaps:']):
                # Skip the technical assessment subheadings but include other content
                if line and not line.endswith(':'):
                    if not line.startswith('*'):  # Only add * if it doesn't already have one
                        formatted_feedback.append(f"* {line}")
                    else:
                        formatted_feedback.append(line)
        
        # Add a blank line between sections
        formatted_feedback.append("")
    
    # Remove the last empty line if present
    if formatted_feedback and not formatted_feedback[-1]:
        formatted_feedback.pop()
        
    return formatted_feedback

def generate_excel_report(json_data, transcript_text, output_file):
    """Generate Excel report based on JSON data and transcript."""
    # Extract information from transcript
    transcript_info = extract_transcript_info(transcript_text)
    
    # Extract categories and scores from JSON
    categories = json_data['categories']
    
    # Find lowest scoring parameter and its shortcomings in each category
    clarification_needed = []
    
    for cat_key, cat_data in categories.items():
        lowest_params = find_lowest_scoring_parameter(cat_data)
        clarification_needed.extend(lowest_params)
    
    # Generate improved areas of improvement from feedback
    improvement_areas = generate_feedback_for_excel(json_data)
    
    # Determine SI No based on existing data
    si_no = 1
    if os.path.exists(output_file):
        try:
            existing_df = pd.read_excel(output_file)
            si_no = len(existing_df) + 1
        except Exception as e:
            print(f"Warning: Could not read existing file for SI No determination: {e}")
    
    # Create DataFrame for new row with modified columns
    df = pd.DataFrame({
        "SI No": [si_no],
        "Vet name": [transcript_info["vet_name"]],
        "Order ID": [transcript_info["order_id"]],
        "Call Start time": [transcript_info["call_start_time"]],
        "Call Duration": [transcript_info["call_duration"]],
        "Achieved score": [json_data["overall"]["percentage_score"]],
        "Parameter needs clarification": ["\n".join([f"• {item}" for item in clarification_needed])],
        "Areas of improvement": ["\n".join(improvement_areas)]
    })
    
    # Function to safely attempt appending to Excel file with multiple retries
    def safe_append_to_excel(df, output_file, attempts=3, delay=1):
        """Try to append data to Excel with multiple attempts and better error handling."""
        for attempt in range(attempts):
            try:
                # Ensure the file exists
                if not os.path.exists(output_file):
                    print(f"File {output_file} doesn't exist, creating new file.")
                    create_new_excel_file(df, output_file)
                    return True
                
                # Load the existing workbook with data_only=True to resolve any formulas
                book = openpyxl.load_workbook(output_file, data_only=True)
                sheet = book.active
                
                # Verify the workbook structure
                required_columns = ["SI No", "Vet name", "Order ID", "Call Start time", 
                                "Call Duration", "Achieved score", 
                                "Parameter needs clarification", "Areas of improvement"]
                
                # Get header row
                header_row = [cell.value for cell in sheet[1]]
                
                # Check if all required columns exist
                missing_columns = [col for col in required_columns if col not in header_row]
                if missing_columns:
                    print(f"Warning: Existing file missing columns: {missing_columns}")
                    print("Creating a new file with correct structure...")
                    # Backup the old file
                    backup_file = f"{output_file}.bak"
                    if os.path.exists(backup_file):
                        os.remove(backup_file)
                    os.rename(output_file, backup_file)
                    print(f"Existing file backed up to {backup_file}")
                    create_new_excel_file(df, output_file)
                    return True
                
                # Write the new data row by row
                row_idx = sheet.max_row + 1  # Start at the next available row
                print(f"Appending at row {row_idx} in {output_file}")
                
                for _, row in df.iterrows():
                    for c_idx, value in enumerate(row):
                        # Find column index matching the DataFrame column
                        col_name = df.columns[c_idx]
                        if col_name in header_row:
                            col_idx = header_row.index(col_name) + 1
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            
                            # Special handling for "Areas of improvement" column
                            if col_name == "Areas of improvement" and value:
                                # Apply formatting for bold headings (Ø headings) and Medicine Name/Dosage
                                lines = str(value).split('\n')
                                formatted_text = ""
                                
                                for i, line in enumerate(lines):
                                    if line.startswith('Ø'):
                                        # For headings, we'll set a bold style later
                                        formatted_text += line + '\n'
                                    else:
                                        formatted_text += line + '\n'
                                
                                # Remove the trailing newline
                                if formatted_text.endswith('\n'):
                                    formatted_text = formatted_text[:-1]
                                
                                # Set the formatted text
                                cell.value = formatted_text
                                
                                # We'll handle styling in a different way instead of using RichText
                            else:
                                cell.value = value
                                
                            cell.alignment = openpyxl.styles.Alignment(vertical='top', horizontal='left', wrap_text=True)
                    # Move to next row after writing all columns
                    row_idx += 1
                
                # Adjust row height for new rows
                for idx in range(sheet.max_row - len(df) + 1, sheet.max_row + 1):
                    # Get maximum number of lines in any cell of this row
                    max_lines = 1
                    for cell in sheet[idx]:
                        if cell.value:
                            lines = str(cell.value).split('\n')
                            if len(lines) > max_lines:
                                max_lines = len(lines)
                    
                    # Set row height based on content lines (minimum 20)
                    row_height = max(20, min(15 * max_lines, 400))  # Cap at 400
                    sheet.row_dimensions[idx].height = row_height
                
                # Save the workbook to a temporary file first
                temp_file = f"{output_file}.temp"
                book.save(temp_file)
                book.close()  # Explicitly close the workbook
                
                # If save succeeded, replace the original file
                if os.path.exists(output_file):
                    os.remove(output_file)
                os.rename(temp_file, output_file)
                
                print(f"Successfully appended data to existing file: {output_file}")
                return True
                
            except PermissionError as e:
                print(f"Attempt {attempt+1}: Permission error - file may be open in another application: {e}")
                if attempt < attempts - 1:
                    print(f"Retrying in {delay} seconds...")
                    import time
                    time.sleep(delay)
                    delay *= 2  # Exponential backoff
                else:
                    print(f"Failed to append after {attempts} attempts due to permissions.")
                    return False
                    
            except Exception as e:
                print(f"Attempt {attempt+1}: Error appending to existing file: {e}")
                import traceback
                traceback.print_exc()
                if attempt < attempts - 1:
                    print(f"Retrying in {delay} seconds...")
                    import time
                    time.sleep(delay)
                    delay *= 2  # Exponential backoff
                else:
                    print(f"Failed to append after {attempts} attempts.")
                    return False
        return False
    
    if os.path.exists(output_file):
        success = safe_append_to_excel(df, output_file)
        if not success:
            print("Creating a new file instead...")
            # Create a timestamped backup of the original file
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = f"{output_file}_{timestamp}.bak"
            try:
                import shutil
                shutil.copy2(output_file, backup_file)
                print(f"Backed up existing file to {backup_file}")
                create_new_excel_file(df, output_file)
            except Exception as e:
                print(f"Error creating backup: {e}")
                # Try creating with a different filename
                new_file = f"veterinary_call_report_{timestamp}.xlsx"
                create_new_excel_file(df, new_file)
                print(f"Created new file with different name: {new_file}")
                return df
    else:
        create_new_excel_file(df, output_file)
    
    return df

def create_new_excel_file(df, output_file):
    """Create a new Excel file with proper formatting."""
    try:
        # Before writing, make sure we can create/write to this file
        test_handle = open(output_file, 'a')
        test_handle.close()
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            # Get the worksheet
            worksheet = writer.sheets['Sheet1']
            
            # Format all cells to be top-left aligned
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(vertical='top', horizontal='left', wrap_text=True)
            
            # Format the "Areas of improvement" column with more basic approach
            for row_idx in range(2, len(df) + 2):  # Start from 2 to skip header row
                # Get the "Areas of improvement" cell
                col_idx = df.columns.get_loc("Areas of improvement") + 1
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                if cell.value:
                    # Apply simple formatting - we'll use string manipulation instead
                    content = str(cell.value)
                    formatted_content = []
                    
                    for line in content.split('\n'):
                        # Just keep the content as it is without trying to apply bold formatting
                        formatted_content.append(line)
                    
                    # Join the lines back together
                    cell.value = '\n'.join(formatted_content)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            # Calculate max length based on content lines
                            lines = str(cell.value).split('\n')
                            for line in lines:
                                if len(line) > max_length:
                                    max_length = len(line)
                    except:
                        pass
                # Add some padding and set width
                adjusted_width = max_length + 2
                # Limit the maximum width
                if adjusted_width > 50:
                    adjusted_width = 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
            # Set row heights for better readability
            for row_idx in range(1, len(df) + 2):  # +2 for header row
                # Calculate height based on content
                max_lines = 1
                for cell in worksheet[row_idx]:
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        if len(lines) > max_lines:
                            max_lines = len(lines)
                
                # Set row height based on content lines (minimum 20)
                row_height = max(20, min(15 * max_lines, 400))  # Cap at 400
                worksheet.row_dimensions[row_idx].height = row_height
                
        print(f"Successfully created new Excel file: {output_file}")
        return True
        
    except PermissionError as pe:
        print(f"Permission error creating file {output_file}: {pe}")
        print("The file may be open in another application or you don't have write permissions.")
        return False
        
    except Exception as e:
        print(f"Error creating Excel file {output_file}: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    parser = argparse.ArgumentParser(description="Generate Excel report and feedback from transcript and JSON data.")
    parser.add_argument("json_file", type=str, help="Path to the JSON file containing score data.")
    parser.add_argument("transcript_file", type=str, help="Path to the transcript text file.")
    parser.add_argument("--output", type=str, default="veterinary_call_report.xlsx", 
                        help="Path to save the Excel report (default: veterinary_call_report.xlsx)")
    args = parser.parse_args()
    
    # Load JSON
    try:
        with open(args.json_file, 'r', encoding='utf-8') as jf:
            json_data = json.load(jf)
    except FileNotFoundError:
        print(f"Error: JSON file '{args.json_file}' not found.")
        return
    except json.JSONDecodeError:
        print(f"Error: Could not parse JSON file '{args.json_file}'.")
        return
        
    # Load Transcript
    try:
        with open(args.transcript_file, 'r', encoding='utf-8') as tf:
            transcript_text = tf.read()
    except FileNotFoundError:
        print(f"Error: Transcript file '{args.transcript_file}' not found.")
        return
        
    # Generate Excel report
    try:
        df = generate_excel_report(json_data, transcript_text, args.output)
        print(f"✅ Excel report saved as '{args.output}'")
        
        # Print feedback
        print("\n📝 Sample feedback:")
        feedback = generate_feedback_for_excel(json_data)
        print("\n".join(feedback))
    except Exception as e:
        print(f"Error generating Excel report: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()